from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta, date
from calendar import monthrange
from decimal import Decimal
import json
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Employee, Department, Designation, Attendance, Leave, Payroll, ProductionOutput, PieceRateSetting, Loan
from .forms import EmployeeForm, AttendanceForm, LeaveForm, PayrollForm, ProductionOutputForm, PieceRateSettingForm, LoanForm, BulkAttendanceForm

# Helper function for role-based access
def is_hr_or_admin(user):
    return user.is_superuser or user.groups.filter(name='HR').exists()


def get_attendance_summary(employee, month, year):
    """Build a day-by-day attendance picture for one employee/month,
    reconciling Attendance records with approved Leave so payroll and
    leave stay in sync:

    - A day with an approved (paid) leave covering it -> paid leave day,
      never counted as absent, even if no Attendance row exists or it
      was previously marked 'absent'.
    - A day with an approved *unpaid* leave covering it -> unpaid leave
      day, which (like an absence) attracts a payroll deduction.
    - A day with no Attendance row and no approved leave -> unexcused
      absence (per company policy: "if no leave is approved, the day
      counts as absent").
    - Future days within the month are not yet counted either way.
    """
    days_in_month = monthrange(year, month)[1]
    first_day = date(year, month, 1)
    last_day = date(year, month, days_in_month)
    today = date.today()
    last_countable_day = min(last_day, today)

    attendance_by_date = {
        a.date: a for a in Attendance.objects.filter(
            employee=employee, date__range=(first_day, last_day)
        )
    }

    leave_by_date = {}
    approved_leaves = Leave.objects.filter(
        employee=employee, status='approved',
        start_date__lte=last_day, end_date__gte=first_day,
    )
    for leave in approved_leaves:
        d = max(leave.start_date, first_day)
        end = min(leave.end_date, last_day)
        while d <= end:
            leave_by_date[d] = leave
            d += timedelta(days=1)

    days_present = 0
    days_late = 0
    days_half_day = 0
    days_paid_leave = 0
    days_unpaid_leave = 0
    days_absent = 0
    total_overtime_hours = Decimal('0')

    d = first_day
    while d <= last_countable_day:
        attendance = attendance_by_date.get(d)
        leave = leave_by_date.get(d)

        if attendance:
            total_overtime_hours += attendance.overtime_hours or 0

        if leave is not None:
            # An approved leave covering this day always wins - that's
            # the leave/payroll sync the policy requires.
            if leave.is_paid:
                days_paid_leave += 1
            else:
                days_unpaid_leave += 1
        elif attendance:
            if attendance.status == 'present':
                days_present += 1
            elif attendance.status == 'late':
                days_present += 1
                days_late += 1
            elif attendance.status == 'half_day':
                days_half_day += 1
            elif attendance.status == 'holiday':
                pass
            elif attendance.status == 'leave':
                # Marked as leave but no matching approved Leave record
                # (e.g. it was later rejected/cancelled) - treat as absent.
                days_absent += 1
            else:  # 'absent'
                days_absent += 1
        else:
            # No attendance recorded and no approved leave -> absent.
            days_absent += 1

        d += timedelta(days=1)

    return {
        'days_in_month': days_in_month,
        'days_present': days_present,
        'days_late': days_late,
        'days_half_day': days_half_day,
        'days_paid_leave': days_paid_leave,
        'days_unpaid_leave': days_unpaid_leave,
        'days_absent': days_absent,
        'deduction_days': days_absent + days_unpaid_leave,
        'total_overtime_hours': total_overtime_hours,
    }


@login_required
def hr_dashboard(request):
    """HR Dashboard Overview"""
    context = {
        'active': 'hr',
        'page_title': 'HR Dashboard',
    }
    
    # Employee Statistics
    context['total_employees'] = Employee.objects.filter(is_active=True).count()
    context['total_departments'] = Department.objects.count()
    context['total_designations'] = Designation.objects.count()
    
    # Employee by category
    context['production_staff'] = Employee.objects.filter(
        is_active=True,
        skill_category__in=['cutting', 'stitching', 'finishing', 'washing', 'qc']
    ).count()
    context['supervisors'] = Employee.objects.filter(skill_category='supervisor', is_active=True).count()
    context['management'] = Employee.objects.filter(skill_category='management', is_active=True).count()
    
    # Gender distribution
    context['male_employees'] = Employee.objects.filter(gender='male', is_active=True).count()
    context['female_employees'] = Employee.objects.filter(gender='female', is_active=True).count()
    
    # Today's attendance
    today = date.today()
    context['today_attendance'] = Attendance.objects.filter(date=today).count()
    context['today_present'] = Attendance.objects.filter(date=today, status='present').count()
    context['today_absent'] = Attendance.objects.filter(date=today, status='absent').count()
    context['today_late'] = Attendance.objects.filter(date=today, is_late=True).count()
    
    # Leave Statistics
    context['pending_leaves'] = Leave.objects.filter(status='pending').count()
    context['approved_leaves'] = Leave.objects.filter(status='approved').count()
    
    # Monthly Payroll Summary
    current_month = date.today().month
    current_year = date.today().year
    payrolls = Payroll.objects.filter(month=current_month, year=current_year, status='processed')
    context['monthly_payroll_total'] = payrolls.aggregate(total=Sum('net_payable'))['total'] or 0
    context['monthly_payroll_count'] = payrolls.count()
    
    # Recent Activities
    context['recent_employees'] = Employee.objects.order_by('-created_at')[:5]
    context['recent_leaves'] = Leave.objects.order_by('-created_at')[:5]
    context['recent_attendances'] = Attendance.objects.order_by('-date')[:5]
    
    # Chart Data - Monthly Hiring
    monthly_hiring = []
    for i in range(6):
        month = date.today().month - i
        year = date.today().year
        if month <= 0:
            month += 12
            year -= 1
        count = Employee.objects.filter(joining_date__month=month, joining_date__year=year).count()
        monthly_hiring.append(count)
    context['monthly_hiring'] = monthly_hiring[::-1]
    context['months'] = [date(date.today().year, m, 1).strftime('%B') for m in range(date.today().month-5, date.today().month+1)]
    
    # Chart Data - Attendance Rate Last 7 Days
    attendance_rate = []
    for i in range(6, -1, -1):
        day = date.today() - timedelta(days=i)
        total = Employee.objects.filter(is_active=True).count()
        present = Attendance.objects.filter(date=day, status='present').count()
        rate = (present / total * 100) if total > 0 else 0
        attendance_rate.append(rate)
    context['attendance_rate'] = attendance_rate
    context['attendance_days'] = [(date.today() - timedelta(days=i)).strftime('%a') for i in range(6, -1, -1)]
    
    return render(request, 'hr/dashboard.html', context)

@login_required
def employee_list(request):
    """List all employees"""
    employees = Employee.objects.select_related('department', 'designation').all()
    
    # Filter by department
    department = request.GET.get('department')
    if department:
        employees = employees.filter(department_id=department)
    
    # Filter by skill
    skill = request.GET.get('skill')
    if skill:
        employees = employees.filter(skill_category=skill)
    
    # Search
    search = request.GET.get('search')
    if search:
        employees = employees.filter(
            Q(full_name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(phone_number__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(employees, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'active': 'hr',
        'page_title': 'Employee List',
        'employees': page_obj,
        'departments': Department.objects.all(),
        'skill_categories': Employee.SKILL_CATEGORIES,
        'current_department': department,
        'current_skill': skill,
        'search': search,
    }
    return render(request, 'hr/employee_list.html', context)

@login_required
def employee_detail(request, pk):
    """Employee details view"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Get attendance for last 30 days
    last_30_days = date.today() - timedelta(days=30)
    recent_attendance = Attendance.objects.filter(employee=employee, date__gte=last_30_days).order_by('-date')[:30]
    
    # Get leaves
    leaves = Leave.objects.filter(employee=employee).order_by('-start_date')[:10]
    
    # Get payroll history
    payrolls = Payroll.objects.filter(employee=employee).order_by('-year', '-month')[:6]
    
    # Get production outputs
    productions = ProductionOutput.objects.filter(employee=employee).order_by('-date')[:10]
    
    context = {
        'active': 'hr',
        'page_title': f'Employee Details - {employee.full_name}',
        'employee': employee,
        'recent_attendance': recent_attendance,
        'leaves': leaves,
        'payrolls': payrolls,
        'productions': productions,
    }
    return render(request, 'hr/employee_detail.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def add_employee(request):
    """Add new employee"""
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.created_by = request.user
            employee.save()
            
            # Create user account if needed
            if request.POST.get('create_user'):
                username = employee.employee_id.lower()
                password = request.POST.get('password', 'password123')
                user = User.objects.create_user(username=username, password=password, email=employee.email)
                employee.user = user
                employee.save()
            
            messages.success(request, f'Employee {employee.full_name} added successfully!')
            return redirect('hr:employee_detail', pk=employee.pk)
    else:
        form = EmployeeForm()
    
    context = {
        'active': 'hr',
        'page_title': 'Add Employee',
        'form': form,
    }
    return render(request, 'hr/employee_form.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def edit_employee(request, pk):
    """Edit employee"""
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f'Employee {employee.full_name} updated successfully!')
            return redirect('hr:employee_detail', pk=employee.pk)
    else:
        form = EmployeeForm(instance=employee)
    
    context = {
        'active': 'hr',
        'page_title': 'Edit Employee',
        'form': form,
        'employee': employee,
    }
    return render(request, 'hr/employee_form.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def delete_employee(request, pk):
    """Delete employee (soft delete)"""
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.is_active = False
        employee.save()
        messages.success(request, f'Employee {employee.full_name} deactivated successfully!')
        return redirect('hr:employee_list')
    
    context = {
        'employee': employee,
    }
    return render(request, 'hr/employee_confirm_delete.html', context)

@login_required
def attendance_view(request):
    """Attendance management"""
    date_filter = request.GET.get('date', date.today())
    if isinstance(date_filter, str):
        date_filter = datetime.strptime(date_filter, '%Y-%m-%d').date()
    
    attendances = Attendance.objects.filter(date=date_filter).select_related('employee')
    
    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():
            attendance = form.save(commit=False)
            attendance.marked_by = request.user
            attendance.save()
            messages.success(request, 'Attendance marked successfully!')
            return redirect('hr:attendance_view')
    else:
        form = AttendanceForm(initial={'date': date_filter})
    
    # Get employees without attendance for this date
    employees_with_attendance = attendances.values_list('employee_id', flat=True)
    employees_without = Employee.objects.filter(is_active=True).exclude(id__in=employees_with_attendance)
    
    context = {
        'active': 'hr',
        'page_title': 'Attendance Management',
        'attendances': attendances,
        'employees_without': employees_without,
        'form': form,
        'current_date': date_filter,
    }
    return render(request, 'hr/attendance.html', context)

@login_required
def bulk_attendance(request):
    """Bulk attendance marking"""
    if request.method == 'POST':
        form = BulkAttendanceForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']
            employees = form.cleaned_data['employees']
            status = form.cleaned_data['status']
            
            for employee in employees:
                attendance, created = Attendance.objects.get_or_create(
                    employee=employee,
                    date=date,
                    defaults={'status': status, 'marked_by': request.user}
                )
                if not created:
                    attendance.status = status
                    attendance.marked_by = request.user
                    attendance.save()
            
            messages.success(request, f'Bulk attendance marked for {employees.count()} employees!')
            return redirect('hr:attendance_view')
    else:
        form = BulkAttendanceForm()
    
    context = {
        'active': 'hr',
        'page_title': 'Bulk Attendance',
        'form': form,
    }
    return render(request, 'hr/bulk_attendance.html', context)

@login_required
def leave_requests(request):
    """Leave request management"""
    leaves = Leave.objects.select_related('employee').all().order_by('-created_at')
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        leaves = leaves.filter(status=status)
    
    context = {
        'active': 'hr',
        'page_title': 'Leave Requests',
        'leaves': leaves,
        'current_status': status,
    }
    return render(request, 'hr/leave_requests.html', context)

@login_required
def apply_leave(request):
    """Apply for leave"""
    if request.method == 'POST':
        form = LeaveForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.save()
            messages.success(request, 'Leave request submitted successfully!')
            return redirect('hr:leave_requests')
    else:
        form = LeaveForm()

    current_year = date.today().year
    employee_balances = {
        str(emp.id): emp.leave_balance(current_year)
        for emp in Employee.objects.filter(is_active=True)
    }

    context = {
        'active': 'hr',
        'page_title': 'Apply for Leave',
        'form': form,
        'employee_balances_json': json.dumps(employee_balances),
        'current_year': current_year,
    }
    return render(request, 'hr/leave_form.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def approve_leave(request, pk):
    """Approve or reject leave request"""
    leave = get_object_or_404(Leave, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            leave.status = 'approved'
            leave.approved_by = request.user
            leave.approved_date = date.today()
            leave.save()

            # Sync attendance for the leave period so payroll never
            # counts these days as unexcused absences.
            tag = f'leave#{leave.pk}'
            d = leave.start_date
            while d <= leave.end_date:
                Attendance.objects.update_or_create(
                    employee=leave.employee, date=d,
                    defaults={
                        'status': 'leave',
                        'remarks': f'On approved {leave.get_leave_type_display()} ({tag})',
                        'marked_by': request.user,
                    }
                )
                d += timedelta(days=1)
            messages.success(request, 'Leave approved successfully! Attendance has been updated for the leave period.')
        elif action == 'reject':
            tag = f'leave#{leave.pk}'
            leave.status = 'rejected'
            leave.save()
            # If attendance was already synced to 'leave' for this
            # request (e.g. previously approved, now being corrected),
            # revert those days back to absent.
            Attendance.objects.filter(
                employee=leave.employee, date__range=(leave.start_date, leave.end_date),
                status='leave', remarks__icontains=f'({tag})'
            ).update(status='absent', remarks='Leave request rejected')
            messages.warning(request, 'Leave rejected!')
        return redirect('hr:leave_requests')
    
    context = {
        'leave': leave,
        'balance': leave.employee.leave_balance(leave.start_date.year).get(leave.leave_type),
    }
    return render(request, 'hr/approve_leave.html', context)

@login_required
def payroll_list(request):
    """Payroll list"""
    payrolls = Payroll.objects.select_related('employee').all().order_by('-year', '-month')
    
    # Filter by month/year
    month = request.GET.get('month')
    year = request.GET.get('year')

    current_year = date.today().year

    oldest_year = Payroll.objects.order_by('year').values_list('year', flat=True).first()

    if oldest_year:
        years = range(current_year, oldest_year - 1, -1)
    else:
        years = [current_year]

    if month and year:
        payrolls = payrolls.filter(month=month, year=year)
    
    context = {
        'active': 'hr',
        'page_title': 'Payroll Management',
        'payrolls': payrolls,
        'months': range(1, 13),
        'current_year': date.today().year,
        'years': years,
    }
    return render(request, 'hr/payroll_list.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def generate_payroll(request):
    """Generate payroll for employees - simple flow:
    1. HR picks month/year.
    2. Every active employee is listed with their default salary
       (basic + house rent + medical + transport, taken straight from
       the Employee record), plus attendance pulled from Attendance/Leave
       records: days present, days on approved (paid) leave, and days
       absent. Any day with no attendance record and no approved leave
       counts as an unexcused absence and is pre-filled as a deduction.
       Overtime (from logged overtime hours) and a bonus field are also
       editable here. HR can edit any of these before confirming.
    3. On confirm, payroll records are created. If the entered base
       salary amount differs from the default, the difference is
       recorded as an incentive (increase) or deduction (decrease) along
       with the note HR wrote explaining why; absence deduction, overtime
       amount, and bonus are stored as entered.
    """
    current_year = date.today().year
    oldest_year = Payroll.objects.order_by('year').values_list('year', flat=True).first()

    if oldest_year:
        years = range(current_year, oldest_year - 1, -1)
    else:
        years = [current_year]

    if request.method == 'POST' and request.POST.get('step') == 'confirm':
        # Step 2: save the (possibly edited) payroll for each employee
        month = int(request.POST.get('month'))
        year = int(request.POST.get('year'))
        employee_ids = request.POST.getlist('employee_id')

        def _decimal(field_name, default):
            raw = request.POST.get(field_name, '').strip()
            try:
                return Decimal(raw) if raw else default
            except Exception:
                return default

        saved_count = 0
        for emp_id in employee_ids:
            employee = get_object_or_404(Employee, pk=emp_id)

            default_salary = (
                employee.basic_salary + employee.house_rent +
                employee.medical_allowance + employee.transport_allowance
            )

            # Recompute attendance/leave for the month fresh from the
            # source records - this is what keeps payroll and leave in
            # sync, rather than trusting whatever was last rendered.
            summary = get_attendance_summary(employee, month, year)
            per_day_rate = (default_salary / summary['days_in_month']) if summary['days_in_month'] else Decimal('0')
            suggested_absence_deduction = (per_day_rate * summary['deduction_days']).quantize(Decimal('0.01'))
            suggested_overtime_amount = (employee.overtime_hourly_rate * summary['total_overtime_hours']).quantize(Decimal('0.01'))

            salary_amount = _decimal(f'salary_{emp_id}', default_salary)
            absence_deduction = _decimal(f'absence_deduction_{emp_id}', suggested_absence_deduction)
            overtime_amount = _decimal(f'overtime_amount_{emp_id}', suggested_overtime_amount)
            bonus = _decimal(f'bonus_{emp_id}', Decimal('0'))

            remarks = request.POST.get(f'remarks_{emp_id}', '').strip()
            difference = salary_amount - default_salary

            incentives = difference if difference > 0 else Decimal('0')
            other_deductions = abs(difference) if difference < 0 else Decimal('0')

            payroll, created = Payroll.objects.update_or_create(
                employee=employee,
                month=month,
                year=year,
                defaults={
                    'basic_pay': employee.basic_salary,
                    'house_rent': employee.house_rent,
                    'medical_allowance': employee.medical_allowance,
                    'transport_allowance': employee.transport_allowance,
                    'incentives': incentives,
                    'other_deductions': other_deductions,
                    'overtime_amount': overtime_amount,
                    'bonus': bonus,
                    'absence_deduction': absence_deduction,
                    'total_working_days': summary['days_in_month'],
                    'days_present': summary['days_present'],
                    'days_absent': summary['deduction_days'],
                    'days_late': summary['days_late'],
                    'total_overtime_hours': summary['total_overtime_hours'],
                    'remarks': remarks,
                    'status': 'processed',
                    'generated_by': request.user,
                }
            )
            payroll.calculate_totals()
            payroll.save()
            saved_count += 1

        messages.success(request, f'Payroll generated for {saved_count} employee(s) for {month}/{year}!')
        return redirect('hr:payroll_list')

    elif request.method == 'POST':
        # Step 1: month/year chosen - show the editable preview
        month = int(request.POST.get('month'))
        year = int(request.POST.get('year'))

        employees = Employee.objects.filter(is_active=True).select_related('department', 'designation')

        employee_rows = []
        for employee in employees:
            default_salary = (
                employee.basic_salary + employee.house_rent +
                employee.medical_allowance + employee.transport_allowance
            )
            summary = get_attendance_summary(employee, month, year)
            per_day_rate = (default_salary / summary['days_in_month']) if summary['days_in_month'] else Decimal('0')
            suggested_absence_deduction = (per_day_rate * summary['deduction_days']).quantize(Decimal('0.01'))
            suggested_overtime_amount = (employee.overtime_hourly_rate * summary['total_overtime_hours']).quantize(Decimal('0.01'))

            employee_rows.append({
                'employee': employee,
                'basic_salary': employee.basic_salary,
                'house_rent': employee.house_rent,
                'medical_allowance': employee.medical_allowance,
                'transport_allowance': employee.transport_allowance,
                'default_salary': default_salary,
                'attendance': summary,
                'absence_deduction': suggested_absence_deduction,
                'overtime_amount': suggested_overtime_amount,
                'bonus': Decimal('0'),
            })

        # Warn if payroll already exists for this month/year (will be overwritten)
        existing_count = Payroll.objects.filter(month=month, year=year).count()

        context = {
            'active': 'hr',
            'page_title': 'Generate Payroll - Review',
            'month': month,
            'year': year,
            'employee_rows': employee_rows,
            'existing_count': existing_count,
        }
        return render(request, 'hr/generate_payroll_preview.html', context)

    # GET: show the month/year selection form
    context = {
        'active': 'hr',
        'page_title': 'Generate Payroll',
        'months': range(1, 13),
        'current_year': date.today().year,
        'years': years,
    }
    return render(request, 'hr/generate_payroll.html', context)
@login_required
@user_passes_test(is_hr_or_admin)
def download_payroll_template(request):
    """Download an Excel template pre-filled with every active employee's
    default salary breakdown, attendance-based absence deduction, and
    suggested overtime for the chosen month/year. HR edits the
    'Salary Amount', 'Absence Deduction', 'Overtime Amount', 'Bonus' and
    'Note' columns and re-uploads the file."""
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')

    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Payroll'

    headers = ['Employee ID', 'Name', 'Department', 'Basic', 'House Rent',
               'Medical', 'Transport', 'Default Salary', 'Days Present',
               'Days Absent (incl. unpaid leave)', 'Days Paid Leave',
               'Suggested Absence Deduction', 'Salary Amount',
               'Absence Deduction', 'Overtime Amount', 'Bonus', 'Note']
    sheet.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    for col_num in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    employees = Employee.objects.filter(is_active=True).select_related('department').order_by('employee_id')
    try:
        month_int = int(month)
        year_int = int(year)
    except (TypeError, ValueError):
        month_int = year_int = None

    for employee in employees:
        default_salary = (
            employee.basic_salary + employee.house_rent +
            employee.medical_allowance + employee.transport_allowance
        )

        if month_int and year_int:
            summary = get_attendance_summary(employee, month_int, year_int)
            per_day_rate = (default_salary / summary['days_in_month']) if summary['days_in_month'] else Decimal('0')
            suggested_absence_deduction = (per_day_rate * summary['deduction_days']).quantize(Decimal('0.01'))
            suggested_overtime_amount = (employee.overtime_hourly_rate * summary['total_overtime_hours']).quantize(Decimal('0.01'))
        else:
            summary = {'days_present': 0, 'deduction_days': 0, 'days_paid_leave': 0}
            suggested_absence_deduction = Decimal('0')
            suggested_overtime_amount = Decimal('0')

        sheet.append([
            employee.employee_id,
            employee.full_name,
            employee.department.name if employee.department else '',
            float(employee.basic_salary),
            float(employee.house_rent),
            float(employee.medical_allowance),
            float(employee.transport_allowance),
            float(default_salary),
            summary['days_present'],
            summary['deduction_days'],
            summary['days_paid_leave'],
            float(suggested_absence_deduction),
            float(default_salary),
            float(suggested_absence_deduction),
            float(suggested_overtime_amount),
            0,
            '',
        ])

    widths = {'A': 14, 'B': 22, 'C': 16, 'D': 10, 'E': 10, 'F': 10, 'G': 10,
              'H': 13, 'I': 11, 'J': 13, 'K': 11, 'L': 13, 'M': 13, 'N': 14,
              'O': 14, 'P': 10, 'Q': 32}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    # Lock the calculated/reference columns so HR can't accidentally edit them,
    # while leaving Salary Amount, Absence Deduction, Overtime Amount,
    # Bonus, and Note open for editing.
    sheet.protection.sheet = True
    sheet.protection.password = ''
    locked_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    editable_cols = ['M', 'N', 'O', 'P', 'Q']
    for row_num in range(1, sheet.max_row + 1):
        for col_letter in locked_cols:
            sheet[f'{col_letter}{row_num}'].protection = sheet[f'{col_letter}{row_num}'].protection.copy(locked=True)
        for col_letter in editable_cols:
            sheet[f'{col_letter}{row_num}'].protection = sheet[f'{col_letter}{row_num}'].protection.copy(locked=False)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'payroll_template_{month}_{year}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_hr_or_admin)
def upload_payroll_excel(request):
    """Read an uploaded, filled-in payroll template and show the same
    review/confirm screen used by the manual flow, pre-populated with the
    amounts and notes from the spreadsheet."""
    if request.method != 'POST':
        return redirect('hr:generate_payroll')

    month = request.POST.get('month')
    year = request.POST.get('year')
    excel_file = request.FILES.get('excel_file')

    if not month or not year:
        messages.error(request, 'Please select month and year before uploading.')
        return redirect('hr:generate_payroll')

    if not excel_file:
        messages.error(request, 'Please choose an Excel file to upload.')
        return redirect('hr:generate_payroll')

    month = int(month)
    year = int(year)

    try:
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active
    except Exception:
        messages.error(request, 'Could not read that file. Please upload a valid .xlsx file (use the downloaded template).')
        return redirect('hr:generate_payroll')

    employee_rows = []
    skipped_ids = []
    seen_ids = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ''):
            continue

        employee_id = str(row[0]).strip()
        if employee_id in seen_ids:
            continue
        seen_ids.add(employee_id)

        employee = Employee.objects.filter(
            employee_id=employee_id, is_active=True
        ).select_related('department').first()

        if not employee:
            skipped_ids.append(employee_id)
            continue

        # Always recompute the default and attendance summary from the
        # live Employee/Attendance/Leave records - never trust the
        # reference/breakdown columns in the uploaded file, only the
        # employee_id and the editable Salary Amount / Absence Deduction /
        # Overtime Amount / Bonus / Note columns.
        default_salary = (
            employee.basic_salary + employee.house_rent +
            employee.medical_allowance + employee.transport_allowance
        )
        summary = get_attendance_summary(employee, month, year)
        per_day_rate = (default_salary / summary['days_in_month']) if summary['days_in_month'] else Decimal('0')
        suggested_absence_deduction = (per_day_rate * summary['deduction_days']).quantize(Decimal('0.01'))
        suggested_overtime_amount = (employee.overtime_hourly_rate * summary['total_overtime_hours']).quantize(Decimal('0.01'))

        def _cell_decimal(index, default):
            value = row[index] if len(row) > index else None
            try:
                return Decimal(str(value)) if value not in (None, '') else default
            except Exception:
                return default

        entered_salary = _cell_decimal(12, default_salary)
        entered_absence_deduction = _cell_decimal(13, suggested_absence_deduction)
        entered_overtime_amount = _cell_decimal(14, suggested_overtime_amount)
        entered_bonus = _cell_decimal(15, Decimal('0'))

        remarks = str(row[16]).strip() if len(row) > 16 and row[16] not in (None, '') else ''

        employee_rows.append({
            'employee': employee,
            'basic_salary': employee.basic_salary,
            'house_rent': employee.house_rent,
            'medical_allowance': employee.medical_allowance,
            'transport_allowance': employee.transport_allowance,
            'default_salary': default_salary,
            'entered_salary': entered_salary,
            'attendance': summary,
            'absence_deduction': entered_absence_deduction,
            'overtime_amount': entered_overtime_amount,
            'bonus': entered_bonus,
            'remarks': remarks,
        })

    if not employee_rows:
        messages.error(request, 'No valid employee rows were found in that file.')
        return redirect('hr:generate_payroll')

    if skipped_ids:
        messages.warning(
            request,
            f"Skipped unknown or inactive employee ID(s): {', '.join(skipped_ids)}"
        )

    existing_count = Payroll.objects.filter(month=month, year=year).count()

    context = {
        'active': 'hr',
        'page_title': 'Generate Payroll - Review',
        'month': month,
        'year': year,
        'employee_rows': employee_rows,
        'existing_count': existing_count,
        'source': 'excel',
    }
    return render(request, 'hr/generate_payroll_preview.html', context)

@login_required
def production_tracking(request):
    """Production output tracking"""
    productions = ProductionOutput.objects.select_related('employee').all().order_by('-date')
    
    # Date filter
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date and to_date:
        productions = productions.filter(date__gte=from_date, date__lte=to_date)
    
    # Style filter
    style = request.GET.get('style')
    if style:
        productions = productions.filter(style_number=style)
    
    # Summary statistics
    total_quantity = productions.aggregate(total=Sum('quantity_produced'))['total'] or 0
    total_defective = productions.aggregate(total=Sum('defective_quantity'))['total'] or 0
    avg_quality = productions.aggregate(avg=Avg('quality_rate'))['avg'] or 0
    
    context = {
        'active': 'hr',
        'page_title': 'Production Tracking',
        'productions': productions,
        'total_quantity': total_quantity,
        'total_defective': total_defective,
        'avg_quality': avg_quality,
        'styles': ProductionOutput.objects.values_list('style_number', flat=True).distinct(),
    }
    return render(request, 'hr/production_tracking.html', context)

@login_required
def export_attendance_csv(request):
    """Export attendance data to CSV"""
    date_filter = request.GET.get('date', date.today())
    if isinstance(date_filter, str):
        date_filter = datetime.strptime(date_filter, '%Y-%m-%d').date()
    
    attendances = Attendance.objects.filter(date=date_filter).select_related('employee')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_{date_filter}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Employee ID', 'Name', 'Department', 'Check In', 'Check Out', 'Status', 'Working Hours', 'Overtime'])
    
    for attendance in attendances:
        writer.writerow([
            attendance.employee.employee_id,
            attendance.employee.full_name,
            attendance.employee.department.name if attendance.employee.department else '',
            attendance.check_in_time,
            attendance.check_out_time,
            attendance.get_status_display(),
            attendance.working_hours,
            attendance.overtime_hours,
        ])
    
    return response

@login_required
def department_list(request):
    """Department management"""
    departments = Department.objects.prefetch_related('designations', 'employees').all()
    
    context = {
        'active': 'hr',
        'page_title': 'Departments',
        'departments': departments,
    }
    return render(request, 'hr/department_list.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def add_department(request):
    """Add new department"""
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        description = request.POST.get('description')
        
        department = Department.objects.create(
            name=name,
            code=code,
            description=description,
            hod=request.POST.get('hod')
        )
        messages.success(request, f'Department {name} added successfully!')
        return redirect('hr:department_list')
    
    context = {
        'active': 'hr',
        'page_title': 'Add Department',
    }
    return render(request, 'hr/department_form.html', context)


@login_required
def designation_list(request):
    """Designation management"""
    designations = Designation.objects.select_related('department').all().order_by('department__name', 'name')
    
    # Filter by department
    department_id = request.GET.get('department')
    if department_id:
        designations = designations.filter(department_id=department_id)
    
    context = {
        'active': 'hr',
        'page_title': 'Designations',
        'designations': designations,
        'departments': Department.objects.all(),
        'selected_department': department_id,
    }
    return render(request, 'hr/designation_list.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def add_designation(request):
    """Add new designation"""
    if request.method == 'POST':
        name = request.POST.get('name')
        department_id = request.POST.get('department')
        grade = request.POST.get('grade')
        basic_salary_min = request.POST.get('basic_salary_min') or 0
        basic_salary_max = request.POST.get('basic_salary_max') or 0
        
        # Validation
        if not name or not department_id:
            messages.error(request, 'Name and Department are required!')
            return redirect('hr:add_designation')
        
        # Check if designation already exists in the same department
        if Designation.objects.filter(name=name, department_id=department_id).exists():
            messages.error(request, f'Designation "{name}" already exists in this department!')
            return redirect('hr:add_designation')
        
        # Validate salary range
        if float(basic_salary_min) > float(basic_salary_max):
            messages.error(request, 'Minimum salary cannot be greater than maximum salary!')
            return redirect('hr:add_designation')
        
        designation = Designation.objects.create(
            name=name,
            department_id=department_id,
            grade=grade,
            basic_salary_min=basic_salary_min,
            basic_salary_max=basic_salary_max
        )
        messages.success(request, f'Designation "{name}" added successfully!')
        return redirect('hr:designation_list')
    
    context = {
        'active': 'hr',
        'page_title': 'Add Designation',
        'departments': Department.objects.all(),
        'is_edit': False,
    }
    return render(request, 'hr/designation_form.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def edit_designation(request, pk):
    """Edit designation"""
    designation = get_object_or_404(Designation, pk=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        department_id = request.POST.get('department')
        grade = request.POST.get('grade')
        basic_salary_min = request.POST.get('basic_salary_min') or 0
        basic_salary_max = request.POST.get('basic_salary_max') or 0
        
        # Validation
        if not name or not department_id:
            messages.error(request, 'Name and Department are required!')
            return redirect('hr:edit_designation', pk=pk)
        
        # Check if designation already exists in the same department (excluding current)
        if Designation.objects.filter(name=name, department_id=department_id).exclude(pk=pk).exists():
            messages.error(request, f'Designation "{name}" already exists in this department!')
            return redirect('hr:edit_designation', pk=pk)
        
        # Validate salary range
        if float(basic_salary_min) > float(basic_salary_max):
            messages.error(request, 'Minimum salary cannot be greater than maximum salary!')
            return redirect('hr:edit_designation', pk=pk)
        
        designation.name = name
        designation.department_id = department_id
        designation.grade = grade
        designation.basic_salary_min = basic_salary_min
        designation.basic_salary_max = basic_salary_max
        designation.save()
        
        messages.success(request, f'Designation "{name}" updated successfully!')
        return redirect('hr:designation_list')
    
    context = {
        'active': 'hr',
        'page_title': 'Edit Designation',
        'designation': designation,
        'departments': Department.objects.all(),
        'is_edit': True,
    }
    return render(request, 'hr/designation_form.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def delete_designation(request, pk):
    """Delete designation"""
    designation = get_object_or_404(Designation, pk=pk)
    
    if request.method == 'POST':
        designation_name = designation.name
        designation.delete()
        messages.success(request, f'Designation "{designation_name}" deleted successfully!')
        return redirect('hr:designation_list')
    
    context = {
        'active': 'hr',
        'page_title': 'Delete Designation',
        'designation': designation,
    }
    return render(request, 'hr/designation_delete_confirm.html', context)

@login_required
def designation_detail(request, pk):
    """View designation details"""
    designation = get_object_or_404(Designation.objects.select_related('department'), pk=pk)
    
    context = {
        'active': 'hr',
        'page_title': 'Designation Details',
        'designation': designation,
    }
    return render(request, 'hr/designation_detail.html', context)